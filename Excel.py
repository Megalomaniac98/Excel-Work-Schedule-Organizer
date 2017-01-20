import openpyxl as Excel
def monthstr(a):
    month = str(a.value)
    for k in range(0,len(month)):
        if month[k] == ' ':
            temp = k
            break
    month = month[0:temp]
    month = month.lower()
    print(month)
    return month

def cellinfo(a):
    temp = str(a)
    for k in range(0,len(temp)):
        if temp[k] == '.':
            temp = temp[k+1:len(temp)]
            break
    lengthtemp = int(len(temp))
    print(lengthtemp)
    print(temp)
    if temp[lengthtemp] == '>':
        temp = temp[k+1:len(temp)-1]

    return temp


def daysthismonth(month):

    if month =='january':
        lastdate = 31
    if month =='february':
        lastdate = 28
    if month =='march':
        lastdate = 31
    if month =='april':
        lastdate = 30
    if month =='may':
        lastdate = 31
    if month =='june':
        lastdate = 30
    if month =='july':
        lastdate = 31
    if month =='august':
        lastdate = 31
    if month =='september':
        lastdate = 30
    if month =='october':
        lastdate = 31
    if month =='november':
        lastdate = 30
    if month =='december':
        lastdate = 31
    return lastdate

def date(a , first, last):
    date = 0
    cell = str(a)
    cellnumber = 0
    for k in range(0,len(cell)):
        if cell[k] == '.':
            cellnumber = cell[k+2:len(cell)-1]
            cellnumber = int(cellnumber)
            break
    if cellnumber >= 3 and cellnumber <= 10:
        if str(day(a)) == 'Monday':
            date = int(first)
        if str(day(a)) == 'Tuesday':
            date =  int(first) + 1
        if str(day(a)) == 'Wednesday':
            date =  int(first) + 2
        if str(day(a)) == 'Thursday':
            date =  int(first) + 3
        if str(day(a)) == 'Friday':
            date =  int(first) + 4
        if str(day(a)) == 'Saturday':
            date =  int(first) + 5
    if cellnumber >= 11 and cellnumber <= 18:
        if str(day(a)) == 'Monday':
            date =  int(first) + 7
        if str(day(a)) == 'Tuesday':
            date =  int(first) + 8
        if str(day(a)) == 'Wednesday':
            date =  int(first) + 9
        if str(day(a)) == 'Thursday':
            date =  int(first) + 10
        if str(day(a)) == 'Friday':
            date =  int(first) + 11
        if str(day(a)) == 'Saturday':
            date =  int(first) + 12
    if cellnumber >= 19 and cellnumber <= 26:
        if str(day(a)) == 'Monday':
            date =  int(first) + 14
        if str(day(a)) == 'Tuesday':
            date =  int(first) + 15
        if str(day(a)) == 'Wednesday':
            date =  int(first) + 16
        if str(day(a)) == 'Thursday':
            date =  int(first) + 17
        if str(day(a)) == 'Friday':
            date =  int(first) + 18
        if str(day(a)) == 'Saturday':
            date =  int(first) + 19
    if cellnumber >= 27 and cellnumber <= 34:
        if str(day(a)) == 'Monday':
            date =  int(first) + 21
        if str(day(a)) == 'Tuesday':
            date =  int(first) + 22
        if str(day(a)) == 'Wednesday':
            date =  int(first) + 23
        if str(day(a)) == 'Thursday':
            date =  int(first) + 24
        if str(day(a)) == 'Friday':
            date =  int(first) + 25
        if str(day(a)) == 'Saturday':
            date =  int(first) + 26
    if cellnumber >= 35 and cellnumber <= 42:
        if str(day(a)) == 'Monday':
            date =  int(first) + 28
        if str(day(a)) == 'Tuesday':
            date =  int(first) + 29
        if str(day(a)) == 'Wednesday':
            date =  int(first) + 30
        if str(day(a)) == 'Thursday':
            date =  int(first) + 31
        if str(day(a)) == 'Friday':
            date =  int(first) + 32
        if str(day(a)) == 'Saturday':
            date =  int(first) + 33
    if date > int(last):
        date -= int(last)
    return date
def day(a):
    cell = str(a)
    celldetail = cellinfo(a)
    temp = celldetail[1:len(celldetail)]
    cell_Letter = celldetail[0]
    print(cell_Letter)
    if cell_Letter == 'B':
        return 'Monday'
    if cell_Letter == 'D':
        return 'Tuesday'
    if cell_Letter == 'F':
        return 'Wednesday'
    if cell_Letter == 'H':
        return 'Thursday'
    if cell_Letter == 'J':
        return 'Friday'
    if cell_Letter == 'L':
        return 'Saturday'
def sumHours(a):
        for k in range(0,len(a)):
            if a[k] == '-':
                first_time = int(a[k-1])
                second_time = int(a[k+1])
            if a[k] == ':':
                semicolon = int(a[k+1])

        if second_time < 7:
            second_time += 12
        if first_time < 7:
            first_time += 12
        sums = second_time - first_time
        if semicolon == 3:
            sums += 0.5
        return sums

def Excel_read():
    wb = Excel.load_workbook('JANUARY DRAFT.xlsx')
    ws = wb.active
    hours_worked = 0
    first_date = input('What is the date of the first Monday on the Calender?')

    month = monthstr(ws.cell('A1'))
    lastday = daysthismonth(month)
    for i in range(1,43):
        for j in range(1,13):
            a = ws.cell(row = i, column = j)

            if a.value != None:
                if isinstance(a.value,str) :
                    for k in range(0,len(a.value)):
                        if a.value[k] == 'A':
                            if a.value[k+1] == 'r' and a.value[k+2] == 'i':
                                length = len(a.value)
                                if a.value[k+4:length] == '':
                                    print('')
                                    print('9:30-6:30pm on ' + str(day(a)) + ' the ' + str(date(a, first_date, lastday))+ ' of '+month)
                                    print('Number of hours worked was: 8')
                                    hours_worked += 8
                                else:
                                    hours = a.value[k+4:length]
                                    print('')
                                    print(str(hours) + 'pm on ' + str(day(a)) + ' the ' + str(date(a, first_date ,lastday)) + ' of '+month)
                                    print('Number of hours worked was: '+str(sumHours(hours)))
                                    hours_worked+= sumHours(hours)
    print('')
    print('Monthly Hours ' + str(hours_worked))
    print('Weekly Hours ' + str(hours_worked/5))

def Main():
    print('---------------------------------------------')
    Excel_read()

Main()
